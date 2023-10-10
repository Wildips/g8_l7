from zipfile import ZipFile
from xlrd import open_workbook
from openpyxl import load_workbook
from pypdf import PdfReader


def test_txt_file(start_stop):
    archive_path, test_data = start_stop
    with ZipFile(archive_path, 'r') as zip_file:
        for row in zip_file.namelist():

            file_extension = row[row.find('.') + len('.'):]

            if file_extension == "txt":
                with open(row, 'r') as fn:
                    assert test_data['txt']['file_text'] in fn.read(), "Текст не эквивалентен"


def test_pdf_file(start_stop):
    archive_path, test_data = start_stop
    with ZipFile(archive_path, 'r') as zip_file:
        for row in zip_file.namelist():

            file_extension = row[row.find('.') + len('.'):]
            if file_extension == "pdf":
                reader = PdfReader(row)
                assert test_data['pdf']['amount_of_sheets'] == len(reader.pages), "Количество страниц не эквивалентно"
                page = reader.pages[1]
                assert test_data['pdf']['page_text'] == page.extract_text(), "Текст на странице не эквивалентен"


def test_xlsx_file(start_stop):
    archive_path, test_data = start_stop
    with ZipFile(archive_path, 'r') as zip_file:
        for row in zip_file.namelist():

            file_extension = row[row.find('.') + len('.'):]

            if file_extension == "xlsx":
                workbook = load_workbook(zip_file.open(row, 'r'))
                sheet = workbook.active
                assert test_data['xlsx']['sheet_crossing'] == sheet.cell(row=3, column=2).value, \
                    "Значение ячеек не эквивалентно"


def test_xls_file(start_stop):
    archive_path, test_data = start_stop
    with ZipFile(archive_path, 'r') as zip_file:
        for row in zip_file.namelist():

            file_extension = row[row.find('.') + len('.'):]

            if file_extension == "xls":
                workbook = open_workbook(row)
                assert test_data['xls']['amount_of_sheets'] == workbook.nsheets, "Количество страниц не эквивалентно"
                assert test_data['xls']['sheets_names_list'] == workbook.sheet_names(), "Имена страниц не эквивалентны"
                # only for current example
                sheet = workbook.sheet_by_index(0)
                assert test_data['xls']['sheet_rows'] == sheet.nrows, "Количество сток на странице не эквивалентно"
                assert test_data['xls'][
                           'sheet_columns'] == sheet.ncols, "Количество колонок на странице не эквивалентно"
                assert test_data['xls']['sheet_crossing'] == sheet.cell_value(9, 2), "Значение ячеек не эквивалентно"
