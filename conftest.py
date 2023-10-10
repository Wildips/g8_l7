import os
import pytest
from zipfile import ZipFile
from xlrd import open_workbook
from openpyxl import load_workbook
from pypdf import PdfReader

PROJECT_PATH = os.path.dirname(os.path.abspath(__file__))
RESOURCES_PATH = os.path.join(PROJECT_PATH, "resources")
TMP_PATH = os.path.join(PROJECT_PATH, "tmp")
FILES_CONTENT_PART = {
    'xls': {'amount_of_sheets': 0, 'sheets_names_list': [], 'sheet_rows': 0, 'sheet_columns': 0, 'sheet_crossing': ''},
    'xlsx': {'sheet_crossing': ''},
    'txt': {'file_text': ''},
    'pdf': {'amount_of_sheets': 0, 'page_text': ''}}


def create_archive():
    resources_list = os.listdir(RESOURCES_PATH)
    if not os.path.exists(TMP_PATH):
        os.mkdir(TMP_PATH)
    with ZipFile(os.path.join(TMP_PATH, 'test_archive.zip'), "w") as zip_file:
        for res_file in resources_list:
            zip_file.write(rf'resources/{res_file}')


def remove_archive():
    if os.path.isfile(os.path.join(TMP_PATH, 'test_archive.zip')):
        os.remove(os.path.join(TMP_PATH, 'test_archive.zip'))


def get_test_data():
    resources_list = os.listdir(RESOURCES_PATH)

    for row in resources_list:
        file_extension = row[row.find('.') + len('.'):]

        if file_extension == "xls":
            workbook = open_workbook(os.path.abspath(rf'resources/{row}'))
            FILES_CONTENT_PART['xls']['amount_of_sheets'] = workbook.nsheets
            FILES_CONTENT_PART['xls']['sheets_names_list'] = workbook.sheet_names()
            # only for current example
            sheet = workbook.sheet_by_index(0)
            FILES_CONTENT_PART['xls']['sheet_rows'] = sheet.nrows
            FILES_CONTENT_PART['xls']['sheet_columns'] = sheet.ncols
            FILES_CONTENT_PART['xls']['sheet_crossing'] = sheet.cell_value(9, 2)
        elif file_extension == "xslx":
            workbook = load_workbook(os.path.abspath(rf'resources/{row}'))
            sheet = workbook.active()
            FILES_CONTENT_PART['xslx']['sheet_crossing'] = sheet.cell(row=3, column=2).value
        elif file_extension == "pdf":
            reader = PdfReader(os.path.abspath(rf'resources/{row}'))
            FILES_CONTENT_PART['pdf']['amount_of_sheets'] = len(reader.pages)
            page = reader.pages[1]
            FILES_CONTENT_PART['pdf']['page_text'] = page.extract_text()
        elif file_extension == "txt":
            with open(os.path.abspath(rf'resources/{row}'), 'r') as fn:
                FILES_CONTENT_PART['txt']['file_text'] = fn.read()


@pytest.fixture(scope='session', autouse=True)
def start_stop():
    get_test_data()
    create_archive()
    yield os.path.join(TMP_PATH, 'test_archive.zip'), FILES_CONTENT_PART
    remove_archive()
